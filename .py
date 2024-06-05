import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl
import pathlib

class App(ctk.CTk):
    def __init__(self):
        super().__init__() 
        self.layout_config()
        self.appearance()
        self.sistema_principal() 

    def layout_config(self):
        self.title("Sistema de Cadastro ao Cliente CAD Manutenções")
        self.geometry("800x600")
        self.configure(bg="#f0f0f0")  

    def appearance(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', "#fff"])
        self.lb_apm.place(x=40, y=470)
        self.opt_apm = ctk.CTkOptionMenu(self, fg_color="#000060", values=["Claro", "Escuro", "Neutro"], command=self.change_appearance)
        self.opt_apm.place(x=40, y=500)

    def sistema_principal(self):
        self.frame = ctk.CTkFrame(self, width=900, height=100, corner_radius=0, fg_color="#000060") 
        self.frame.place(x=0, y=0)
        
        tittle = ctk.CTkLabel(self.frame, text="Cadastro de Serviços - CAD Manutenções", font=("Helvetica", 24), text_color="#fff")
        tittle.place(x=50, y=10) 
        
        tittle2 = ctk.CTkLabel(self.frame, text="Por favor, preencha todos os campos!", font=("Century Gothic bold", 16), text_color="#fff")
        tittle2.place(x=50, y=50) 

        def submit():
            bloco = pathlib.Path("Clientes.xlsx")
            if bloco.is_file():
                pass
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws['A1'] = "Nome Completo"
                ws['B1'] = "Contato"
                ws['C1'] = "Endereço"
                ws['D1'] = "Problema"
                ws['E1'] = "Obs"
                wb.save("Clientes.xlsx")

            # Pegando Entrys
            name = nome_value.get()
            contato = contato_value.get()
            endereco = endereco_value.get()
            problema = problema_value.get()
            obs = self.obs_entry.get(0.0, END)
            
            bloco = openpyxl.load_workbook('Clientes.xlsx')
            folha = bloco.active
            folha.append([name, contato, endereco, problema, obs])
            bloco.save("Clientes.xlsx")
            messagebox.showinfo("Sistema", "Dados salvos com sucesso, agora só aguardar")
            

        def clear():
            nome_value.set("")
            contato_value.set("")
            endereco_value.set("")
            problema_value.set("")
            self.obs_entry.delete(0.0, END)

        # Variáveis de texto
        nome_value = StringVar()
        contato_value = StringVar()
        endereco_value = StringVar()
        problema_value = StringVar()

        # Entrys
        self.nome_entry = ctk.CTkEntry(self, width=350, textvariable=nome_value, font=("Century Gothic", 16), fg_color="#000")
        self.contato_entry = ctk.CTkEntry(self, width=350, textvariable=contato_value, font=("Century Gothic", 16), fg_color="#000")
        self.endereco_entry = ctk.CTkEntry(self, width=350, textvariable=endereco_value, font=("Century Gothic", 16), fg_color="#000")
        self.problema_entry = ctk.CTkEntry(self, width=350, textvariable=problema_value, font=("Century Gothic", 16), fg_color="#000")

        # Labels
        nome_label = ctk.CTkLabel(self, text="Nome:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        nome_label.place(x=50, y=120)
        contato_label = ctk.CTkLabel(self, text="Contato:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        contato_label.place(x=50, y=160)
        endereco_label = ctk.CTkLabel(self, text="Endereço:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        endereco_label.place(x=50, y=200)
        problema_label = ctk.CTkLabel(self, text="Problema:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        problema_label.place(x=50, y=240)
        obs_label = ctk.CTkLabel(self, text="Observações:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        obs_label.place(x=50, y=280)

        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131")
        btn_submit.place(x=300, y=500)

        btn_clear = ctk.CTkButton(self, text="Limpar".upper(), command=clear, fg_color="#555", hover_color="#333")
        btn_clear.place(x=500, y=500)

        # Colocar Entrys na tela
        self.nome_entry.place(x=150, y=120)
        self.contato_entry.place(x=150, y=160)
        self.endereco_entry.place(x=150, y=200) 
        self.problema_entry.place(x=150, y=240)

        # Entrada de observações
        self.obs_entry = ctk.CTkTextbox(self, width=500, height=150, font=("Arial", 18), border_color="#aaa", border_width=2, fg_color="#000")
        self.obs_entry.place(x=150, y=280)

    def change_appearance(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

if __name__ == "__main__":
    app = App()
    app.mainloop()